from locale import format_string
from os import path, rename
from time import sleep
import yaml as ym
import time
import regex as re
from xlrd import Book as xlrdbook
from openpyxl import workbook as openpybook


jira_key_regex = r"[A-Z]+\-\d+(?=\s)"
param_regex = r"[a-zA-Z]+_\w*"
req_code_regex = r"[A-Z]+_\w*"


def print_elapsed_time(elapsed_time, timeout):

    time_res = min(0.25, timeout / 10)

    if (
        (elapsed_time == 0)
        or (not hasattr(print_elapsed_time, "back_str"))
        or (print_elapsed_time.back_str is None)
    ):
        print_elapsed_time.back_str = ""
    print(print_elapsed_time.back_str, end="")
    sleep(time_res)
    elapsed_time_str = f"{elapsed_time:6.2f}/{timeout:3.2f}s "
    print(elapsed_time_str, end="", flush=True)
    print_elapsed_time.back_str = "\b" * len(elapsed_time_str)


def set_test_params(tst, motor_id):
    """setup the test parameters

    Args:
        tst ([type]): [description]
    """

    motor_no = tst[motor_id]["axis_no"]
    gate_index = 0 if motor_no < 5 else 1
    chan_index = motor_no - gate_index * 4 - 1
    comp_mot_no = motor_no + 8
    enctable_no = motor_no

    if "L1" not in tst[motor_id]:
        tst[motor_id]["L1"] = motor_no

    if "L2" not in tst[motor_id]:
        tst[motor_id]["L2"] = gate_index

    if "L3" not in tst[motor_id]:
        tst[motor_id]["L3"] = chan_index

    if "L4" not in tst[motor_id]:
        tst[motor_id]["L4"] = tst[motor_id]["L2"]

    if "L5" not in tst[motor_id]:
        tst[motor_id]["L5"] = tst[motor_id]["L3"]

    if "L6" not in tst[motor_id]:
        tst[motor_id]["L6"] = motor_no - 1

    if "L7" not in tst[motor_id]:
        tst[motor_id]["L7"] = comp_mot_no

    if "L8" not in tst[motor_id]:
        tst[motor_id]["L8"] = enctable_no

    if "L9" not in tst[motor_id]:
        tst[motor_id]["L9"] = comp_mot_no

    if not "egu" in tst[motor_id]:
        tst[motor_id]["egu"] = "MotU"

    if not "cnt_per_rev" in tst[motor_id]:
        tst[motor_id]["cnt_per_rev"] = (
            tst[motor_id]["fullsteps_per_rev"] * tst[motor_id]["micro_steps"]
        )
    if not "amim" in tst[motor_id]:
        tst[motor_id]["amim"] = (
            tst[motor_id]["egu_per_rev"] / tst[motor_id]["cnt_per_rev"]
        )

    if not "egu_per_mu" in tst[motor_id]:
        tst[motor_id]["egu_per_mu"] = tst[motor_id]["amim"]

    egu_per_mu = tst[motor_id]["egu_per_mu"]

    enc_res = tst[motor_id]["enc_egu_per_mu"] = (
        tst[motor_id]["encoder_scf"]
        if "encoder_scf" in tst[motor_id]
        else (
            tst[motor_id]["encoder_res_egu"]
            if "encoder_res_egu" in tst[motor_id]
            else egu_per_mu
        )
    )

    if all(["smalljog_egu", "bigjog_egu"]) in tst[motor_id]:
        tst[motor_id]["smalljog_steps"] = tst[motor_id]["smalljog_egu"] / egu_per_mu
        tst[motor_id]["bigjog_steps"] = tst[motor_id]["bigjog_egu"] / egu_per_mu
        tst[motor_id]["jog_step_ratio"] = (
            tst[motor_id]["bigjog_egu"] / tst[motor_id]["smalljog_egu"]
        )

    if "HomeOffset_EGU" in [motor_id]:
        tst[motor_id]["HomeOffset"] = tst[motor_id]["HomeOffset_EGU"] / egu_per_mu
        if "attackpos_egu" in tst[motor_id]:
            tst[motor_id]["attackpos_enc"] = (
                tst[motor_id]["attackpos_egu"] + tst[motor_id]["HomeOffset_EGU"]
            ) / enc_res

    if "JogSpeed_EGU" in tst[motor_id]:
        tst[motor_id]["JogSpeed"] = tst[motor_id]["JogSpeed_EGU"] / egu_per_mu / 1000

    if "HomeVel_EGU" in tst[motor_id]:
        tst[motor_id]["HomeVel"] = tst[motor_id]["HomeVel_EGU"] / egu_per_mu / 1000

    if "travel_range_egu" in tst[motor_id]:
        tst[motor_id]["fullrange_steps"] = (
            tst[motor_id]["travel_range_egu"] / egu_per_mu
        )

    if "clearance_egu" in tst[motor_id]:
        tst[motor_id]["clearance_enc"] = tst["clearance_egu"] / enc_res

    return tst

    # it is possible to use multiple gpascii channels,
    # but we don't have a reason to do so, yet!


class ShortHand:
    """
    This class auto completes shorthanded text messages/naes/references based on the recent activity.
    Based on the preset format, it tries to guess
    and complete shorthanded leading ( and maybe trailing ) text.
    """

    # expression = r"\([^\(]*<([^<]*)>[^\(]*\)"
    # expression = re.compile(expression)

    # reversed = re.sub(expression, partial(_group_replacer, data), string)
    short_text = ...  # type: str

    def __init__(
        self, group_formats: list, ditto_char="/", pre_dittos=False, post_dittos=False
    ) -> None:

        self.format_list = group_formats
        self.ditto_char = ditto_char
        self.pre_dittos = pre_dittos
        self.post_dittos = post_dittos

        self.full_expression = ""
        for group in self.format_list:
            self.full_expression += "(" + group[0] + ")*" + group[1]

        self.template = re.compile(self.full_expression)
        self.text_groups = [None] * len(group_formats)

    def long(self, short_text: str):
        """generates long form from short hand input
            by filling in the blanks using latest inputs

        Args:
            short_text (str): [description]

        Returns:
            [type]: [complete long form inferred from history]
        """

        self.short_text = short_text
        self._decompose()
        return self._compose()

    def _decompose(self):

        # parse the new text input:
        match = re.search(self.full_expression, self.short_text)
        match_group = match.group()
        match_groups = re.findall(self.full_expression, self.short_text)
        for mg in match_groups:
            if any(mg):
                match_group = list(mg)
                # the first non-empty group is accepted
                break

        pre_text = True
        post_text = False
        chars_found_count = 0
        dittos_found_count = 0
        for i, old_text in enumerate(self.text_groups):

            if match_group[i]:
                pre_text = False
                chars_found_count += len(match_group[i])
                self.text_groups[i] = match_group[i]
            elif old_text:
                post_text = not pre_text
                expecting_dittos = (self.pre_dittos and pre_text) or (
                    self.post_dittos and post_text
                )
                expected_position = dittos_found_count + chars_found_count

                if (len(self.short_text) > expected_position) and (
                    self.short_text[expected_position] == self.ditto_char
                ):
                    # a missing word... is there a ditto here
                    dittos_found_count += 1
                else:
                    # blank in this placeholder
                    if expecting_dittos:
                        # this is an error, because we haven't found no text and no dittos
                        raise RuntimeError(
                            f"Dittos missing at position {expected_position} of {self.short_text}"
                        )

            else:
                # both are blank, we are confused!!
                raise RuntimeError(
                    f"No pretext to complete {self.short_text}, pretext is {self.text_groups} "
                )

    def _compose(self):
        # compose the long form
        long_text = ""  # type: str
        # compose the full length output from self.text_fields
        for i, group in enumerate(self.format_list):
            long_text += self.text_groups[i] + group[1]

        return long_text


def xlrd_sheet_to_dict(wb, sheet_name, heading_row, first_col, key_col):
    """makes a dict for sheet specified by sheet_name
    keys shall be in rows, (one column)

    excel workbook (.xls or .xlsx) into object wb


    Args:
        wb (xlrd.Book): workbook source object
        sheet_name (str): sheet name
        heading_row (int, optional): where heading row is found (Excel row 1 is index 0). Defaults to 2.
        first_col (int, optional): leftmost column of the table (Excel column 'A' is index 0). Defaults to 2.
        key_col (int, optional): dict keys are picked at thios column (Excel column 'A' is index 0). Defaults to None.

    Returns:
        [dict]: [description]
    """

    assert isinstance(wb, xlrdbook)

    sheet = wb.sheet_by_name(sheet_name)
    sheet_dict = dict()
    for row in range(heading_row + 1, sheet.nrows):

        row_dict = dict()
        for col in range(first_col, sheet.ncols):
            value = sheet.cell(row, col).value
            key = str(sheet.cell(heading_row, col).value).strip()
            try:
                value = str(value).strip()
            except ValueError:
                pass

            if col == key_col:
                req_key = value
                continue

            row_dict.update({key: value})

        sheet_dict.update({req_key: row_dict})

    return sheet_dict


def opxl_sheet_to_dict(wb, sheet_name, heading_row=2, first_col=2, key_col=None):

    assert isinstance(wb, openpybook)
    sheet = wb[sheet_name]
    reqs = dict()
    for row in range(heading_row + 1, sheet.max_row):

        req = dict()
        for col in range(first_col, sheet.max_column):
            value = sheet.cell(row + 1, col + 1).value
            key = str(sheet.cell(heading_row + 1, col + 1).value).strip()
            try:
                value = str((value)).strip()
            except ValueError:
                pass

            if col == key_col:
                req_key = value
                continue

            req.update({key: value})

        reqs.update({req_key: req})

    return reqs


def avoid_overwrite(filepath):

    n_copies = 0
    while path.exists(filepath):
        name, ext = path.splitext(filepath)
        modif_time_str = time.strftime(
            "%y%m%d_%H%M",
            time.localtime(path.getmtime(filepath)),
        )
        n_copies_str = f"({n_copies})" if n_copies > 0 else ""
        try:
            rename(
                filepath,
                f"{name}_{modif_time_str}{n_copies_str}{ext}",
            )
        except FileExistsError:
            # next copy... never overwrite
            n_copies += 1

    return filepath


def parse_params(text):
    """parse text and return parameters in ccc_ccc format split via ","

    Args:
        text ([str]): [text which includes params]

    Returns:
        [list]: [of strings]
    """

    params_list = text if isinstance(text, list) else re.findall(param_regex, text)

    return params_list if len(params_list) > 0 else []


def query_field(any_table, parameters_field):
    """
    finds parameters listed in field "parameters_field" in table "any_table"
    returns the collection of all paameters in parameter_set
    also replaces the original parameter text with parsed list
    """
    parameter_set = set()

    for _record, _fields in any_table.items():
        _fields: dict
        new_parameters = parse_params(_fields[parameters_field])

        parameter_set = parameter_set.union(set(new_parameters))

        # DONE remove the side effect
        # _fields[parameters_field] = new_parameters

    return parameter_set


def dump_obj(myobj, yaml_dump_file):

    with open(yaml_dump_file, "w+") as f:
        try:
            ym.safe_dump(myobj, f, default_flow_style=False)
            success = True
        except Exception:
            success = False
        finally:
            f.close()
            return success


def undump_obj(obj_str, yaml_dump_path):

    with open(path.join(yaml_dump_path, obj_str + ".yaml"), "r") as f:
        myobjects = ym.safe_load(f)
        f.close()

    return myobjects


def dump_class_objs(yourself, obj_str_list, yaml_dump_path):
    """dumps all dicts in the class to corresponding yaml files in requested folder
    yaml file names follow member names

    Args:
        yourself (class): []
        obj_str_list (str): list of members to be dumped
        yaml_dump_path (str): path of dump folder

    Returns:
        [type]: [description]
    """

    list_of_errors = []
    for obj_str in obj_str_list:

        # yaml doesn't know how to represent objects. so to dump dicts of objects, only dump

        if not dump_obj(
            eval("yourself." + obj_str),
            path.join(yaml_dump_path, obj_str + ".yaml"),
        ):
            list_of_errors.append(obj_str)
    return list_of_errors


def time_stamp(filename):

    create_time_str = time.strftime("%y%m%d_%H%M", time.localtime())
    name, ext = path.splitext(filename)

    return f"{name}_{create_time_str}{ext}"


def isPmacNumber(s: str):

    if s.startswith("$"):
        # check if it is a valid hex
        try:
            int(s[1:], 16)
        except ValueError:
            return False
        else:
            return True
    else:
        # check if it is a valid decimal
        return (
            s.replace("e-", "")
            .replace("e+", "")
            .lstrip("+-")
            .replace(".", "", 1)
            .isdigit()
        )


if __name__ == "__main__":
    print_elapsed_time.back_str = None
    pass
